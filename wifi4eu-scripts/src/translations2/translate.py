from openpyxl import load_workbook
from collections import OrderedDict
import simplejson as json
wb = load_workbook(filename = 'translations.xlsx')
ws = wb['To translate']

languages = []
keys = []
translations_list = []

for idx, col in enumerate(ws.iter_cols(min_col=2, max_col=(ws.max_column), min_row=1, max_row=1)):
	for cell in col:
		languages.append(cell.value.lower())

for idx, row in enumerate(ws.iter_rows(min_col=1, max_col=1, min_row=2, max_row=(ws.max_row))):
	for cell in row:
		keys.append(cell.value)

for idxCol, col in enumerate(ws.iter_cols(min_col=2, max_col=(ws.max_column), min_row=2, max_row=(ws.max_row))):
	translations = OrderedDict()
	for idxCell, cell in enumerate(col):
		translations[keys[idxCell]] = cell.value
	translations_list.append(translations)
	j = json.dumps(translations, ensure_ascii=False).encode('utf8')
	filename = languages[idxCol] + '.json'
	if languages[idxCol] == 'en (reviewed)':
		filename = 'en.json'
	with open(filename, 'wb') as f:
	    f.write(j)